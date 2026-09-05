"""
Excel report exporter using openpyxl with fallback formatted CSV/HTML for offline flexibility.
Generates multi-worksheet workbooks with summary cards, computation steps, rules applied, and disclaimers.
"""
import os
import csv
from typing import Dict, Any

def export_to_excel(calc_data: Dict[str, Any], output_path: str) -> str:
    """Exports calculation or comparison results to an Excel (.xlsx) file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws_summary = wb.active
        ws_summary.title = "Tax Computation Summary"

        # Theme styling
        header_font = Font(name="Calibri", size=14, bold=True, color="1E3A8A")
        sub_font = Font(name="Calibri", size=11, bold=True, color="1E293B")
        normal_font = Font(name="Calibri", size=11, color="334155")
        bold_font = Font(name="Calibri", size=11, bold=True, color="0F172A")
        fill_header = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        fill_highlight = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        ws_summary["A1"] = "INDIAN INCOME-TAX COMPUTATION REPORT"
        ws_summary["A1"].font = header_font
        ws_summary["A2"] = f"Generated On: {calc_data.get('timestamp', '')}"
        ws_summary["A2"].font = normal_font

        # Taxpayer Details
        row = 4
        ws_summary.cell(row=row, column=1, value="Taxpayer Profile").font = sub_font
        ws_summary.cell(row=row, column=1).fill = fill_header
        row += 1

        profile_fields = [
            ("Taxpayer Name", calc_data.get("taxpayer_name", "N/A")),
            ("Masked PAN", calc_data.get("pan_masked", "N/A")),
            ("Governing Act", calc_data.get("act_name", "Income-tax Act, 1961")),
            ("Tax Regime", calc_data.get("regime", "New Regime")),
            ("Financial Year (FY)", calc_data.get("financial_year", "2024-25")),
            ("Assessment Year (AY)", calc_data.get("assessment_year", "2025-26")),
        ]

        for k, v in profile_fields:
            c1 = ws_summary.cell(row=row, column=1, value=k)
            c2 = ws_summary.cell(row=row, column=2, value=str(v))
            c1.font = normal_font
            c2.font = bold_font
            c1.border = thin_border
            c2.border = thin_border
            row += 1

        # Summary Computation
        row += 1
        ws_summary.cell(row=row, column=1, value="Core Computation").font = sub_font
        ws_summary.cell(row=row, column=1).fill = fill_header
        ws_summary.cell(row=row, column=2, value="Amount (INR)").font = sub_font
        ws_summary.cell(row=row, column=2).fill = fill_header
        row += 1

        summary = calc_data.get("summary", {})
        comp_items = [
            ("Gross Total Income (GTI)", summary.get("gross_total_income", "0")),
            ("Less: Deductions & Exemptions", summary.get("total_deductions", "0")),
            ("Taxable Total Income", summary.get("taxable_total_income", "0")),
            ("Tax at Normal Slab Rates", summary.get("normal_slab_tax", "0")),
            ("Tax at Special Rates (111A / 112A)", summary.get("special_rate_tax", "0")),
            ("Total Tax before Rebate", summary.get("tax_before_rebate", "0")),
            ("Less: Rebate u/s 87A", summary.get("rebate_amount", "0")),
            ("Tax after Rebate", summary.get("tax_after_rebate", "0")),
            ("Add: Surcharge", summary.get("surcharge_amount", "0")),
            ("Add: Health & Education Cess (4%)", summary.get("cess_amount", "0")),
            ("Total Tax Liability", summary.get("total_tax_liability", "0")),
            ("Less: Prepaid Taxes (TDS / TCS / Adv. Tax)", summary.get("total_prepaid_tax", "0")),
            ("Net Balance Payable / (Refund)", summary.get("balance_tax_payable", "0")),
        ]

        for label, val in comp_items:
            c1 = ws_summary.cell(row=row, column=1, value=label)
            c2 = ws_summary.cell(row=row, column=2, value=val)
            c1.font = bold_font if "Total" in label or "Net" in label else normal_font
            c2.font = bold_font if "Total" in label or "Net" in label else normal_font
            if "Net" in label:
                c1.fill = fill_highlight
                c2.fill = fill_highlight
            c1.border = thin_border
            c2.border = thin_border
            row += 1

        # Advance tax schedule sheet
        instalments = calc_data.get("instalment_schedule", [])
        if instalments:
            ws_inst = wb.create_sheet(title="Advance Tax Schedule")
            ws_inst["A1"] = "ADVANCE TAX INSTALMENT DUE DATES & PROGRESSION"
            ws_inst["A1"].font = header_font

            headers = ["Instalment No.", "Statutory Due Date", "Cumulative %", "Required Cumulative Tax (INR)", "Description"]
            for col_idx, h in enumerate(headers, 1):
                cell = ws_inst.cell(row=3, column=col_idx, value=h)
                cell.font = sub_font
                cell.fill = fill_header
                cell.border = thin_border

            inst_row = 4
            for inst in instalments:
                ws_inst.cell(row=inst_row, column=1, value=inst.get("instalment_number", "")).border = thin_border
                ws_inst.cell(row=inst_row, column=2, value=inst.get("due_date", "")).border = thin_border
                ws_inst.cell(row=inst_row, column=3, value=f"{inst.get('cumulative_percent', '')}%").border = thin_border
                ws_inst.cell(row=inst_row, column=4, value=inst.get("required_cumulative_amount", "")).border = thin_border
                ws_inst.cell(row=inst_row, column=5, value=inst.get("description", "")).border = thin_border
                inst_row += 1

        # Audit Trail & Disclaimer Sheet
        ws_audit = wb.create_sheet(title="Audit Trail & Disclaimers")
        ws_audit["A1"] = "CALCULATION AUDIT TRAIL & VERIFIED RULES APPLIED"
        ws_audit["A1"].font = header_font

        audit_row = 3
        ws_audit.cell(row=audit_row, column=1, value="Rules Applied").font = sub_font
        audit_row += 1
        for r in calc_data.get("rules_applied", []):
            ws_audit.cell(row=audit_row, column=1, value=f"• {r}").font = normal_font
            audit_row += 1

        audit_row += 1
        ws_audit.cell(row=audit_row, column=1, value="Step-by-Step Calculation Trail").font = sub_font
        audit_row += 1
        for step in calc_data.get("audit_trail", []):
            ws_audit.cell(row=audit_row, column=1, value=f"• {step}").font = normal_font
            audit_row += 1

        audit_row += 2
        ws_audit.cell(row=audit_row, column=1, value="LEGAL SAFEGUARD DISCLAIMER").font = sub_font
        audit_row += 1
        ws_audit.cell(row=audit_row, column=1, value=calc_data.get("disclaimer", "")).font = normal_font

        # Auto-adjust column widths
        for ws in wb.worksheets:
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

        wb.save(output_path)
        return output_path

    except ImportError:
        # Fallback to CSV if openpyxl is not installed in the environment
        csv_path = output_path.replace(".xlsx", ".csv")
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["INDIAN INCOME-TAX COMPUTATION REPORT"])
            writer.writerow(["Generated On", calc_data.get("timestamp", "")])
            writer.writerow([])
            writer.writerow(["TAX PAYER PROFILE"])
            writer.writerow(["Name", calc_data.get("taxpayer_name", "")])
            writer.writerow(["PAN (Masked)", calc_data.get("pan_masked", "")])
            writer.writerow(["Act", calc_data.get("act_name", "")])
            writer.writerow(["Regime", calc_data.get("regime", "")])
            writer.writerow(["Financial Year", calc_data.get("financial_year", "")])
            writer.writerow([])
            writer.writerow(["COMPUTATION SUMMARY"])
            for k, v in calc_data.get("summary", {}).items():
                writer.writerow([k, v])
            writer.writerow([])
            writer.writerow(["LEGAL DISCLAIMER"])
            writer.writerow([calc_data.get("disclaimer", "")])
        return csv_path
